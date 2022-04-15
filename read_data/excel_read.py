from openpyxl import load_workbook

from web_keys.keys import Key
from read_data.excel_config import excel_style


def get_params(value):
    if value:
        temp = value.split(";")
        data2 = {}
        for i in temp:
            lst = i.split("=", maxsplit=1)
            data2[lst[0]] = lst[1]
        return data2


def run(file_name, log):
    try:
        wb = load_workbook(file_name)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if not sheet["A1"].value:
                continue
            log.info("*********************{}*******************".format(sheet_name))
            for row in sheet:
                if type(row[0].value) != int:
                    continue
                log.info("现在执行的步骤是{}:{}".format(row[1].value, row[3].value))
                data = get_params(row[2].value)
                if row[1].value == "open_browser":
                    key = Key(data["browser_type"], log)
                elif "assert" in row[1].value:
                    res = getattr(key, row[1].value)(**data)
                    cell = row[4]
                    excel_style(res, cell)
                    wb.save(file_name)
                else:
                    if data:
                        getattr(key, row[1].value)(**data)
                    else:
                        getattr(key, row[1].value)()
    except Exception as e:
        log.info("运行异常：{}".format(e))
    finally:
        wb.close()
