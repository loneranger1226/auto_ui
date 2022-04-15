from openpyxl.styles import PatternFill, Font


def excel_style(res, cell):
    if res:
        cell.value = "Pass"
        cell.fill = PatternFill("solid", fgColor="AACF91")
        cell.font = Font(bold=True)
    else:
        cell.value = "Failed"
        cell.fill = PatternFill("solid", fgColor="FF0000")
        cell.font = Font(bold=True)
